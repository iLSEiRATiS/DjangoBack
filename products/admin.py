from decimal import Decimal
from urllib.parse import urlparse
from os.path import basename

import openpyxl
import requests
from django.contrib import admin, messages
from django.core.files.base import ContentFile
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils.text import slugify

from .models import Product, Category, Offer


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ("nombre", "slug", "parent")
    prepopulated_fields = {"slug": ("nombre",)}
    search_fields = ("nombre", "descripcion")
    list_filter = ("parent",)


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ("nombre", "slug", "precio", "user", "categoria", "stock", "activo", "creado_en")
    search_fields = ("nombre", "descripcion", "slug")
    list_filter = ("creado_en", "activo", "categoria")
    list_editable = ("precio", "stock", "activo")
    change_list_template = "admin/products/product/change_list.html"

    product_headers = [
        "sku", "parent_sku", "nombre", "slug", "descripcion", "categoria", "subcategoria", "marca",
        "precio", "costo", "moneda", "stock", "activo", "opcion_1_nombre", "opcion_1_valor",
        "opcion_2_nombre", "opcion_2_valor", "imagen_1", "imagen_2", "meta_title",
        "meta_description", "peso", "largo", "ancho", "alto", "es_destacado", "requiere_envio",
        "gestion_stock",
    ]

    sample_rows = [
        {
            "sku": "REM-BAS-001",
            "parent_sku": "",
            "nombre": "Remera basica",
            "slug": "remera-basica",
            "descripcion": "Remera algodon lisa",
            "categoria": "Ropa",
            "subcategoria": "Remeras",
            "marca": "Acme",
            "precio": 8999,
            "costo": 4500,
            "moneda": "ARS",
            "stock": 120,
            "activo": True,
            "opcion_1_nombre": "",
            "opcion_1_valor": "",
            "opcion_2_nombre": "",
            "opcion_2_valor": "",
            "imagen_1": "https://example.com/rem-bas-001.jpg",
            "imagen_2": "",
            "meta_title": "Remera basica",
            "meta_description": "Remera de algodon basica",
            "peso": 0.3,
            "largo": 30,
            "ancho": 25,
            "alto": 2,
            "es_destacado": True,
            "requiere_envio": True,
            "gestion_stock": True,
        },
        {
            "sku": "REM-BAS-001-M-NEGRO",
            "parent_sku": "REM-BAS-001",
            "nombre": "Remera basica M Negro",
            "slug": "remera-basica-m-negro",
            "descripcion": "Remera algodon talla M color negro",
            "categoria": "Ropa",
            "subcategoria": "Remeras",
            "marca": "Acme",
            "precio": 8999,
            "costo": 4500,
            "moneda": "ARS",
            "stock": 40,
            "activo": True,
            "opcion_1_nombre": "Talle",
            "opcion_1_valor": "M",
            "opcion_2_nombre": "Color",
            "opcion_2_valor": "Negro",
            "imagen_1": "https://example.com/rem-bas-001-m-negro.jpg",
            "imagen_2": "",
            "meta_title": "Remera basica M negro",
            "meta_description": "Remera negra basica talla M",
            "peso": 0.3,
            "largo": 30,
            "ancho": 25,
            "alto": 2,
            "es_destacado": False,
            "requiere_envio": True,
            "gestion_stock": True,
        },
    ]

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "importar-xlsx/",
                self.admin_site.admin_view(self.import_xlsx_view),
                name="products_product_import_xlsx",
            ),
        ]
        return custom + urls

    def _parse_bool(self, value, default=True):
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return value
        s = str(value).strip().lower()
        return s in {"true", "1", "si", "sí", "yes", "y"}

    def _parse_decimal(self, value):
        if value is None or value == "":
            return None
        s = str(value).replace("$", "").replace(" ", "").replace(",", ".")
        try:
            return Decimal(s)
        except Exception:
            return None

    def _parse_int(self, value):
        if value in (None, ""):
            return None
        try:
            return int(float(value))
        except Exception:
            return None

    def _build_slug(self, base):
        candidate = slugify(base or "")[:110] or "producto"
        original = candidate
        i = 1
        while Product.objects.filter(slug=candidate).exists():
            i += 1
            candidate = f"{original}-{i}"
        return candidate

    def _export_workbook(self, rows, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        ws.append(self.product_headers)
        for row in rows:
            ws.append([row.get(h, "") for h in self.product_headers])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def import_xlsx_view(self, request):
        if request.method == "GET" and request.GET.get("sample"):
            return self._export_workbook(self.sample_rows, "productos_ejemplo.xlsx")
        if request.method == "GET" and request.GET.get("template"):
            empty_row = {h: "" for h in self.product_headers}
            return self._export_workbook([empty_row], "plantilla_productos.xlsx")

        created = 0
        updated = 0
        errors = []

        if request.method == "POST":
            upload = request.FILES.get("file")
            if not upload:
                messages.error(request, "Selecciona un archivo XLSX.")
                return redirect(
                    reverse("admin:products_product_import_xlsx")
                )
            try:
                wb = openpyxl.load_workbook(upload, data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    raise ValueError("El archivo está vacío.")
                headers = [str(h or "").strip() for h in rows[0]]
                missing = {"sku", "nombre", "precio"} - set(h.lower() for h in headers)
                if missing:
                    errors.append(f"Faltan columnas obligatorias: {', '.join(sorted(missing))}")
                header_map = {h.strip().lower(): idx for idx, h in enumerate(headers)}
                with transaction.atomic():
                    for idx, raw in enumerate(rows[1:], start=2):
                        data = {h: raw[header_map[h]] if h in header_map and header_map[h] < len(raw) else "" for h in header_map}
                        # normaliza claves originales (con y sin mayúsculas)
                        row_data = {k: data.get(k) for k in header_map}
                        if all(v in ("", None) for v in row_data.values()):
                            continue
                        nombre = row_data.get("nombre") or ""
                        precio = self._parse_decimal(row_data.get("precio"))
                        if not nombre or precio is None:
                            errors.append(f"Fila {idx}: nombre y precio son obligatorios.")
                            continue
                        slug_src = row_data.get("slug") or row_data.get("sku") or nombre
                        slug = slugify(slug_src)[:110] or self._build_slug(nombre)
                        categoria_nombre = row_data.get("categoria") or ""
                        categoria_obj = None
                        if categoria_nombre:
                            categoria_obj, _ = Category.objects.get_or_create(nombre=categoria_nombre)
                        existing = Product.objects.filter(slug=slug).first()
                        is_new = existing is None
                        product = existing or Product(slug=slug, user=request.user)
                        product.nombre = nombre
                        product.descripcion = row_data.get("descripcion") or ""
                        product.precio = precio
                        stock = self._parse_int(row_data.get("stock"))
                        product.stock = stock if stock is not None else 0
                        product.activo = self._parse_bool(row_data.get("activo"), default=True)
                        product.categoria = categoria_obj
                        image_url = row_data.get("imagen_1") or ""
                        if image_url and str(image_url).startswith(("http://", "https://")):
                            try:
                                resp = requests.get(str(image_url), stream=True, timeout=8)
                                if resp.status_code == 200:
                                    filename = basename(urlparse(str(image_url)).path) or f"{slug}.jpg"
                                    product.imagen.save(filename, ContentFile(resp.content), save=False)
                                else:
                                    errors.append(f"Fila {idx}: no se pudo descargar imagen ({resp.status_code}).")
                            except Exception as exc:
                                errors.append(f"Fila {idx}: error descargando imagen ({exc}).")
                        if is_new and not product.slug:
                            product.slug = self._build_slug(nombre)
                        product.save()
                        if is_new:
                            created += 1
                        else:
                            updated += 1
                if created or updated:
                    messages.success(
                        request,
                        f"Importación completada. Nuevos: {created} | Actualizados: {updated}",
                    )
                if errors:
                    for err in errors:
                        messages.error(request, err)
            except Exception as exc:  # pragma: no cover
                messages.error(request, f"No se pudo procesar el XLSX: {exc}")
                return redirect(reverse("admin:products_product_import_xlsx"))

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Importar productos via XLSX",
            "headers": self.product_headers,
            "example_url": f"{reverse('admin:products_product_import_xlsx')}?sample=1",
            "template_url": f"{reverse('admin:products_product_import_xlsx')}?template=1",
            "created": created,
            "updated": updated,
            "errors": errors,
        }
        return TemplateResponse(request, "admin/products/product/import_xlsx.html", context)


@admin.register(Offer)
class OfferAdmin(admin.ModelAdmin):
    list_display = ("nombre", "porcentaje", "producto", "categoria", "activo", "empieza", "termina")
    list_filter = ("activo", "empieza", "termina", "categoria")
    search_fields = ("nombre", "descripcion", "producto__nombre", "categoria__nombre")
    prepopulated_fields = {"slug": ("nombre",)}
    list_editable = ("activo",)
    actions = ["activar_ofertas", "desactivar_ofertas"]

    @admin.action(description="Activar ofertas seleccionadas")
    def activar_ofertas(self, request, queryset):
        queryset.update(activo=True)

    @admin.action(description="Desactivar ofertas seleccionadas")
    def desactivar_ofertas(self, request, queryset):
        queryset.update(activo=False)
